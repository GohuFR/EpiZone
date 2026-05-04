"""
EpiZone — Loader generique

Lit un fichier Excel selon la configuration YAML de la maladie,
et produit un DataFrame normalise de periodes.

Schema de sortie (DataFrame "periodes") :
    code_insee    str    Code INSEE 5 chars (ou marqueur dept si expansion requise)
    commune       str    Nom de la commune
    departement   str    Nom du departement (ou None)
    region        str    Nom de la region (ou None)
    date_debut    date   Date d'entree en zone
    date_fin      date   Date de sortie (NaT si toujours active)
    zone          str    Identifiant de zone (ex: ZP, ZS, ZR, ZR+ZV)
    _is_dept      bool   True si l'entree represente un departement entier
    _dept_code    str    Code departement pour expansion (ex: "36")
"""

from __future__ import annotations

import re
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from .config import DiseaseConfig, SheetConfig, DerivedZoneConfig


# ── Normalisation du code INSEE ──────────────────────────────────────────────

def normalize_code_insee(code: str | int | float | None) -> str | None:
    """
    Normalise un code INSEE :
    - Codes Corse (2A/2B) : gardes tels quels
    - Codes numeriques : paddes sur 5 chiffres
    - Codes departement (XXX, DEPT_) : gardes tels quels (marques ensuite)
    - None/NaN : retourne None
    """
    if code is None or (isinstance(code, float) and np.isnan(code)):
        return None

    code_str = str(code).strip()

    if not code_str:
        return None

    # Marqueurs de departement entier — ne pas toucher
    if "XXX" in code_str or code_str.startswith("DEPT_"):
        return code_str

    # Codes Corse
    if re.match(r"^2[AB]", code_str):
        return code_str

    # Code numerique : padder sur 5 chiffres
    try:
        return f"{int(float(code_str)):05d}"
    except (ValueError, OverflowError):
        return code_str


# ── Lecture d'un onglet Excel ────────────────────────────────────────────────

def _read_sheet(
    excel_path: Path,
    sheet_cfg: SheetConfig,
    disease_cfg: DiseaseConfig,
) -> pd.DataFrame:
    """Lit un onglet Excel et retourne un DataFrame brut normalise."""

    # Lecture avec openpyxl (data_only=True pour resoudre les formules cachees)
    # Note : data_only ne recalcule pas les formules, il lit les valeurs mises
    # en cache par Excel. Si le fichier n'a jamais ete ouvert dans Excel,
    # certaines cellules avec formules seront None.
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb[sheet_cfg.sheet_name]

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 2:
        return pd.DataFrame()

    headers = [str(h) if h else f"_col{i}" for i, h in enumerate(rows[0])]
    df = pd.DataFrame(rows[1:], columns=headers)

    # Supprimer les lignes entierement vides
    df = df.dropna(how="all")

    cols = sheet_cfg.columns

    # ── Extraction des colonnes standard ──────────────────────────────
    result = pd.DataFrame()

    result["code_insee_raw"] = df[cols.code_insee]
    result["commune"] = df.get(cols.commune, pd.Series(dtype=str))
    result["departement"] = df.get(cols.departement, pd.Series(dtype=str)) if cols.departement else None
    result["region"] = df.get(cols.region, pd.Series(dtype=str)) if cols.region else None

    # Dates
    result["date_debut"] = pd.to_datetime(df[cols.date_debut], errors="coerce")
    result["date_fin"] = pd.to_datetime(df[cols.date_fin], errors="coerce")

    # ── Zone ──────────────────────────────────────────────────────────
    if sheet_cfg.zone_id is not None:
        # Zone fixe pour tout l'onglet
        result["zone"] = sheet_cfg.zone_id
    elif sheet_cfg.zone_column:
        # Zone lue depuis une colonne
        raw_zone = df[sheet_cfg.zone_column].astype(str)
        if sheet_cfg.zone_value_map:
            result["zone"] = raw_zone.map(sheet_cfg.zone_value_map)
        else:
            result["zone"] = raw_zone
    else:
        raise ValueError(
            f"Onglet {sheet_cfg.sheet_name}: ni zone_id ni zone_column defini"
        )

    # ── Colonnes supplementaires (ex: raison, raison_fin) ─────────────
    for key, col_name in sheet_cfg.extra_columns.items():
        if col_name in df.columns:
            result[key] = df[col_name]

    # ── Dept num (pour expansion) ─────────────────────────────────────
    if cols.dept_num and cols.dept_num in df.columns:
        result["_dept_num"] = df[cols.dept_num]

    # ── Filtre de lignes (ex: exclure raison=="ZVI" de l'onglet ZS) ──
    if sheet_cfg.row_filter:
        filter_col = sheet_cfg.row_filter.column
        if filter_col in df.columns:
            mask = ~df[filter_col].astype(str).isin(
                sheet_cfg.row_filter.exclude_values
            )
            result = result[mask.values]

    # ── Normalisation code INSEE ──────────────────────────────────────
    result["code_insee"] = result["code_insee_raw"].apply(normalize_code_insee)

    # Supprimer les lignes sans code INSEE ou sans date de debut
    result = result.dropna(subset=["code_insee", "date_debut"])

    # ── Detection des entrees departement-entier ──────────────────────
    result["_is_dept"] = False
    result["_dept_code"] = None

    if disease_cfg.dept_expansion.enabled:
        pattern = disease_cfg.dept_expansion.compiled_pattern
        if pattern:
            for idx, code in result["code_insee"].items():
                m = pattern.match(str(code))
                if m:
                    result.at[idx, "_is_dept"] = True
                    dept_str = m.group(1)
                    result.at[idx, "_dept_code"] = f"{int(dept_str):02d}"

    return result


# ── Construction des zones derivees ──────────────────────────────────────────

def _build_derived_zones(
    excel_path: Path,
    derived_cfgs: list[DerivedZoneConfig],
) -> pd.DataFrame:
    """Construit les periodes des zones derivees (ZVII, ZVI, etc.)."""

    all_derived = []

    for dcfg in derived_cfgs:
        # Lire l'onglet source en brut
        wb = load_workbook(excel_path, read_only=True, data_only=True)
        ws = wb[dcfg.source_sheet]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if len(rows) < 2:
            continue

        headers = [str(h) if h else f"_col{i}" for i, h in enumerate(rows[0])]
        df = pd.DataFrame(rows[1:], columns=headers)
        df = df.dropna(how="all")

        # Appliquer le filtre
        filter_col = dcfg.filter["column"]
        filter_val = dcfg.filter["equals"]
        mask = df[filter_col].astype(str).str.strip() == filter_val
        filtered = df[mask].copy()

        if filtered.empty:
            continue

        # Construire le DataFrame de periodes
        result = pd.DataFrame()
        for key, col_name in dcfg.copy_columns.items():
            result[key] = filtered[col_name].values

        # Code INSEE normalise
        result["code_insee"] = result["code_insee"].apply(normalize_code_insee)

        # Date de debut
        result["date_debut"] = pd.to_datetime(
            filtered[dcfg.date_debut_from].values, errors="coerce"
        )

        # Date de fin
        if dcfg.date_fin_rule:
            # Regle conditionnelle (ex: DNC ZVI)
            rule = dcfg.date_fin_rule
            cond_col = filtered[rule["if_column"]].astype(str).str.strip()
            cond_val = rule["equals"]
            raw_fin = pd.to_datetime(
                filtered[rule["else_from"]].values, errors="coerce"
            )
            result["date_fin"] = pd.NaT
            result.loc[cond_col.values != cond_val, "date_fin"] = raw_fin[
                cond_col.values != cond_val
            ]
        elif dcfg.date_fin_value is None:
            result["date_fin"] = pd.NaT
        else:
            result["date_fin"] = pd.to_datetime(dcfg.date_fin_value)

        result["zone"] = dcfg.id
        result["_is_dept"] = False
        result["_dept_code"] = None

        all_derived.append(result)

    if all_derived:
        return pd.concat(all_derived, ignore_index=True)
    return pd.DataFrame()


# ── Point d'entree principal ─────────────────────────────────────────────────

PERIODES_COLUMNS = [
    "code_insee", "commune", "departement", "region",
    "date_debut", "date_fin", "zone", "_is_dept", "_dept_code",
]


def load_disease_data(
    config: DiseaseConfig,
    data_dir: str | Path,
) -> pd.DataFrame:
    """
    Charge les donnees d'une maladie et retourne un DataFrame normalise
    de periodes, pret pour le moteur de snapshots.

    Parametres
    ----------
    config : DiseaseConfig
        Configuration chargee depuis le YAML.
    data_dir : str | Path
        Repertoire contenant le fichier Excel.

    Retour
    ------
    pd.DataFrame avec les colonnes PERIODES_COLUMNS.
    """
    data_dir = Path(data_dir)
    excel_path = data_dir / config.excel_file

    if not excel_path.exists():
        raise FileNotFoundError(f"Fichier Excel introuvable : {excel_path}")

    print(f"Chargement de {config.name} depuis {excel_path.name}...")

    # ── 1. Lecture des onglets configures ─────────────────────────────
    sheet_dfs = []
    for sheet_cfg in config.sheets:
        print(f"  Onglet '{sheet_cfg.sheet_name}'... ", end="")
        df = _read_sheet(excel_path, sheet_cfg, config)
        print(f"{len(df)} lignes")
        sheet_dfs.append(df)

    # ── 2. Zones derivees ─────────────────────────────────────────────
    if config.derived_zones:
        print(f"  Zones derivees... ", end="")
        derived_df = _build_derived_zones(excel_path, config.derived_zones)
        if not derived_df.empty:
            print(f"{len(derived_df)} lignes")
            sheet_dfs.append(derived_df)
        else:
            print("aucune")

    # ── 3. Assemblage ─────────────────────────────────────────────────
    if not sheet_dfs:
        print("  ⚠ Aucune donnee chargee")
        return pd.DataFrame(columns=PERIODES_COLUMNS)

    periodes = pd.concat(sheet_dfs, ignore_index=True)

    # S'assurer que toutes les colonnes existent
    for col in PERIODES_COLUMNS:
        if col not in periodes.columns:
            periodes[col] = None

    # Selectionner et ordonner
    periodes = periodes[PERIODES_COLUMNS].copy()

    # Nettoyage final
    periodes = periodes.dropna(subset=["code_insee", "date_debut"])
    periodes = periodes.sort_values(["code_insee", "zone", "date_debut"])
    periodes = periodes.reset_index(drop=True)

    # ── Stats ─────────────────────────────────────────────────────────
    n_communes = periodes.loc[~periodes["_is_dept"], "code_insee"].nunique()
    n_dept = periodes.loc[periodes["_is_dept"] == True, "_dept_code"].nunique()
    zones_found = sorted(periodes["zone"].dropna().unique())

    print(f"  → {len(periodes)} periodes totales")
    print(f"  → {n_communes} communes distinctes", end="")
    if n_dept > 0:
        print(f" + {n_dept} departements entiers a expander", end="")
    print()
    print(f"  → Zones : {', '.join(zones_found)}")

    dates = periodes["date_debut"].dropna()
    if not dates.empty:
        print(f"  → Periode : {dates.min().date()} → ", end="")
        fins = periodes["date_fin"].dropna()
        if not fins.empty:
            print(f"{fins.max().date()}")
        else:
            print("en cours")

    return periodes
