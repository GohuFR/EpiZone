"""
EpiZone — Assistant d'import de nouvelles maladies

Analyse un fichier Excel, detecte sa structure,
genere la configuration YAML et declenche le rechargement de l'app.
"""

from __future__ import annotations

import os
import re
import shutil
import time
from pathlib import Path

import openpyxl
import pandas as pd


DEFAULT_COLORS = [
    "#D32F2F", "#E65100", "#F9A825", "#4CAF50",
    "#1976D2", "#8E24AA", "#64B5F6", "#00897B",
]

COLUMN_HINTS = {
    "code_insee": ["code insee", "insee", "code_insee", "code commune"],
    "commune": ["commune", "nom", "lib", "nom_commune"],
    "departement": ["departement", "département", "dept", "dep"],
    "region": ["region", "région"],
    "date_debut": ["date de début", "date de debut", "date_debut", "debut", "début"],
    "date_fin": ["date de fin", "date_fin", "fin"],
    "zone": ["zone", "type zone", "type_zone"],
}


def analyze_excel(filepath: str | Path) -> dict:
    """Analyse un fichier Excel et retourne sa structure."""
    filepath = Path(filepath)
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

    result = {"filename": filepath.name, "sheets": []}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(min_row=1, max_row=min(ws.max_row or 1, 10),
                                  values_only=True))
        if not rows:
            continue

        headers = [str(h).strip() if h else f"Col_{i}" for i, h in enumerate(rows[0])]
        n_rows = (ws.max_row or 1) - 1

        col_info = []
        for j, header in enumerate(headers):
            values = [rows[i][j] for i in range(1, len(rows)) if i < len(rows)]
            sample = next((v for v in values if v is not None), None)
            col_info.append({
                "index": j, "name": header,
                "sample": str(sample)[:50] if sample else "",
            })

        mapping = _auto_detect_columns(headers)

        zone_values = []
        if mapping.get("zone") is not None:
            zone_col_idx = mapping["zone"]
            all_rows = list(ws.iter_rows(min_row=2, values_only=True))
            zone_vals = set()
            for row in all_rows:
                if zone_col_idx < len(row) and row[zone_col_idx]:
                    zone_vals.add(str(row[zone_col_idx]).strip())
            zone_values = sorted(zone_vals)

        result["sheets"].append({
            "name": sheet_name, "n_rows": n_rows,
            "columns": col_info, "headers": headers,
            "mapping": mapping, "zone_values": zone_values,
        })

    wb.close()
    return result


def _auto_detect_columns(headers: list[str]) -> dict:
    mapping = {}
    headers_lower = [h.lower().strip() for h in headers]

    for field, keywords in COLUMN_HINTS.items():
        best_idx, best_score = None, 0
        for i, h in enumerate(headers_lower):
            for kw in keywords:
                if kw == h:
                    best_idx, best_score = i, 100
                    break
                elif kw in h and len(kw) / len(h) > best_score:
                    best_idx = i
                    best_score = len(kw) / len(h)
            if best_score == 100:
                break
        if best_idx is not None and best_score > 0.3:
            mapping[field] = best_idx

    return mapping


def _quote(val: str) -> str:
    """Met en guillemets une valeur YAML de maniere securisee."""
    if not val:
        return '""'
    # Toujours quoter pour eviter les problemes YAML
    val = val.replace('"', '\\"')
    return f'"{val}"'


def generate_config(
    disease_id: str,
    disease_name: str,
    excel_filename: str,
    sheets_config: list[dict],
    accent_color: str = "#E65100",
    map_center: list[float] = None,
    map_zoom: int = 6,
) -> str:
    """Genere le YAML manuellement avec un controle total sur le formatage."""
    if map_center is None:
        map_center = [46.5, 2.5]

    lines = []
    lines.append(f"id: {_quote(disease_id)}")
    lines.append(f"name: {_quote(disease_name)}")
    lines.append(f'subtitle: ""')
    lines.append(f'description: ""')
    lines.append(f"accent_color: {_quote(accent_color)}")
    lines.append(f"excel_file: {_quote(excel_filename)}")
    lines.append("")

    # Zones
    lines.append("zones:")
    seen = set()
    zone_idx = 0
    for sc in sheets_config:
        zid = sc.get("zone_id", sc["sheet_name"])
        if zid and zid not in seen:
            color = sc.get("color", DEFAULT_COLORS[zone_idx % len(DEFAULT_COLORS)])
            label = sc.get("label", zid)
            lines.append(f"  - id: {_quote(zid)}")
            lines.append(f"    label: {_quote(label)}")
            lines.append(f"    color: {_quote(color)}")
            lines.append(f"    priority: {zone_idx + 1}")
            seen.add(zid)
            zone_idx += 1
    lines.append("")

    # Sheets
    lines.append("sheets:")
    for sc in sheets_config:
        sname = sc["sheet_name"]
        zid = sc.get("zone_id", sname)
        cols = sc["columns"]

        lines.append(f"  - sheet_name: {_quote(sname)}")
        lines.append(f"    zone_id: {_quote(zid)}")
        lines.append(f"    columns:")
        lines.append(f"      code_insee: {_quote(cols.get('code_insee', ''))}")
        lines.append(f"      date_debut: {_quote(cols.get('date_debut', ''))}")

        if cols.get("date_fin"):
            lines.append(f"      date_fin: {_quote(cols['date_fin'])}")
        else:
            lines.append(f'      date_fin: "Date de fin"')

        if cols.get("commune"):
            lines.append(f"      commune: {_quote(cols['commune'])}")
        if cols.get("departement"):
            lines.append(f"      departement: {_quote(cols['departement'])}")
        if cols.get("region"):
            lines.append(f"      region: {_quote(cols['region'])}")

    lines.append("")

    # Sections fixes
    lines.append("dept_expansion:")
    lines.append("  enabled: false")
    lines.append("")
    lines.append("derived_zones: []")
    lines.append("combo_zones: []")
    lines.append("")
    lines.append("map:")
    lines.append(f"  center: [{map_center[0]}, {map_center[1]}]")
    lines.append(f"  zoom: {map_zoom}")
    lines.append("")
    lines.append("regulatory:")
    lines.append('  arrete: ""')
    lines.append('  note: ""')
    lines.append("")

    return "\n".join(lines)


def save_import(
    yaml_content: str,
    disease_id: str,
    uploaded_filepath: Path,
    config_dir: Path = Path("configs"),
    data_dir: Path = Path("data"),
) -> dict:
    """Sauvegarde la config YAML et copie le fichier Excel."""
    config_dir.mkdir(parents=True, exist_ok=True)
    data_dir.mkdir(parents=True, exist_ok=True)

    yaml_path = config_dir / f"{disease_id}.yaml"
    excel_dest = data_dir / uploaded_filepath.name

    with open(yaml_path, "w", encoding="utf-8") as f:
        f.write(yaml_content)

    if uploaded_filepath != excel_dest:
        shutil.copy2(uploaded_filepath, excel_dest)

    return {
        "yaml_path": str(yaml_path),
        "excel_path": str(excel_dest),
        "disease_id": disease_id,
    }


def trigger_reload():
    """
    Declenche le rechargement de l'app en mode debug.
    En mode debug, Dash surveille les fichiers .py et redémarre
    automatiquement quand un fichier change.
    """
    # Touch app.py pour declencher le auto-reload de Dash
    app_path = Path("app.py")
    if app_path.exists():
        app_path.touch()
        return True
    return False
